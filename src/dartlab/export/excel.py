"""Company → Excel (.xlsx) 내보내기.

Company의 모든 property를 동적으로 스캔하여 DataFrame이면 시트로 생성.
finance 시계열(IS/BS/CF)은 연도별 피벗, 나머지는 원본 DataFrame 그대로.

사용법::

    from dartlab import Company
    c = Company("005930")
    c.toExcel()                          # 005930_삼성전자.xlsx (전체)
    c.toExcel("output.xlsx")             # 지정 경로
    c.toExcel(modules=["IS", "BS"])      # 시트 선택
    c.toExcel(modules=["dividend", "audit", "employee"])  # 원하는 모듈만

    # 템플릿 기반 내보내기
    from dartlab.export.template import PRESETS
    exportWithTemplate(c, PRESETS["summary"])
    exportWithTemplate(c, PRESETS["governance"], "output.xlsx")
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Optional

import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
	from dartlab.company import Company
	from dartlab.export.template import ExcelTemplate, SheetSpec


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_ACCOUNT_FONT = Font(bold=True, size=10)
_NEGATIVE_FMT = '#,##0;[Red]-#,##0'
_NUMBER_FMT = '#,##0'

_SJ_LABELS = {"IS": "손익계산서", "BS": "재무상태표", "CF": "현금흐름표"}

_ACCOUNT_LABELS = {
	"revenue": "매출액",
	"cost_of_sales": "매출원가",
	"gross_profit": "매출총이익",
	"selling_and_administrative_expenses": "판관비",
	"operating_income": "영업이익",
	"other_income": "기타수익",
	"other_expenses": "기타비용",
	"finance_income": "금융수익",
	"finance_cost": "금융비용",
	"profit_before_tax": "법인세차감전이익",
	"income_tax_expense": "법인세비용",
	"net_income": "당기순이익",
	"comprehensive_income": "총포괄이익",
	"current_assets": "유동자산",
	"cash_and_equivalents": "현금및현금성자산",
	"shortterm_financial_instruments": "단기금융상품",
	"trade_receivables": "매출채권",
	"inventories": "재고자산",
	"non_current_assets": "비유동자산",
	"ppe": "유형자산",
	"intangible_assets": "무형자산",
	"total_assets": "자산총계",
	"current_liabilities": "유동부채",
	"non_current_liabilities": "비유동부채",
	"total_liabilities": "부채총계",
	"total_equity": "자본총계(지배)",
	"equity_nci": "비지배지분",
	"retained_earnings": "이익잉여금",
	"issued_capital": "자본금",
	"operating_cashflow": "영업활동CF",
	"investing_cashflow": "투자활동CF",
	"financing_cashflow": "재무활동CF",
}

_FINANCE_SHEETS = frozenset({"IS", "BS", "CF"})
_SPECIAL_SHEETS = frozenset({"ratios"})


def _autoWidth(ws, minWidth: int = 12, maxWidth: int = 30) -> None:
	for col in ws.columns:
		colLetter = get_column_letter(col[0].column)
		best = minWidth
		for cell in col:
			if cell.value is not None:
				length = len(str(cell.value))
				if isinstance(cell.value, (int, float)):
					length = max(length, 14)
				if length > best:
					best = length
		ws.column_dimensions[colLetter].width = min(best + 2, maxWidth)


def _writeHeaders(ws, headers: list[str]) -> None:
	for colIdx, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=colIdx, value=header)
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.alignment = _HEADER_ALIGN


def _writeFinanceSheet(
	wb: Workbook,
	sjDiv: str,
	series: dict[str, dict[str, list[Optional[float]]]],
	years: list[str],
	*,
	label: str = "",
	columns: list[str] | None = None,
	yearFilter: list[str] | None = None,
) -> None:
	sheetLabel = label or _SJ_LABELS.get(sjDiv, sjDiv)
	ws = wb.create_sheet(title=sheetLabel[:31])

	stmtData = series.get(sjDiv, {})
	if not stmtData:
		return

	displayYears = years
	yearIndices: list[int] | None = None
	if yearFilter:
		yearIndices = [i for i, y in enumerate(years) if y in yearFilter]
		displayYears = [years[i] for i in yearIndices]

	_writeHeaders(ws, ["계정"] + displayYears)

	row = 2
	for snakeId, vals in stmtData.items():
		if columns and snakeId not in columns:
			continue
		effectiveVals = [vals[i] for i in yearIndices] if yearIndices else vals
		if not any(v is not None for v in effectiveVals):
			continue
		ws.cell(row=row, column=1, value=_ACCOUNT_LABELS.get(snakeId, snakeId)).font = _ACCOUNT_FONT
		for colIdx, val in enumerate(effectiveVals, 2):
			if val is not None:
				cell = ws.cell(row=row, column=colIdx, value=round(val))
				cell.number_format = _NEGATIVE_FMT
		row += 1

	_autoWidth(ws)
	ws.freeze_panes = "B2"


def _writeRatiosSheet(wb: Workbook, c: Company, *, label: str = "") -> None:
	from dartlab.engines.common.finance.ratios import calcRatios

	ts = c._cache.get("_finance_q_CFS")
	if ts is None:
		return
	series, _ = ts
	ratios = calcRatios(series)

	ws = wb.create_sheet(title=(label or "재무비율")[:31])
	_writeHeaders(ws, ["지표", "값", "단위"])

	rows = [
		("영업이익률", ratios.operatingMargin, "%"),
		("순이익률", ratios.netMargin, "%"),
		("ROE", ratios.roe, "%"),
		("ROA", ratios.roa, "%"),
		("부채비율", ratios.debtRatio, "%"),
		("유동비율", ratios.currentRatio, "%"),
		("자기자본비율", ratios.equityRatio, "%"),
		("순부채비율", ratios.netDebtRatio, "%"),
		("매출 3Y CAGR", ratios.revenueGrowth3Y, "%"),
		("FCF (TTM)", ratios.fcf, "원"),
		("매출 (TTM)", ratios.revenueTTM, "원"),
		("영업이익 (TTM)", ratios.operatingIncomeTTM, "원"),
		("순이익 (TTM)", ratios.netIncomeTTM, "원"),
	]

	for rowIdx, (rowLabel, val, unit) in enumerate(rows, 2):
		ws.cell(row=rowIdx, column=1, value=rowLabel).font = _ACCOUNT_FONT
		if val is not None:
			if unit == "%":
				ws.cell(row=rowIdx, column=2, value=round(val, 2))
				ws.cell(row=rowIdx, column=2).number_format = '0.00'
			else:
				ws.cell(row=rowIdx, column=2, value=round(val))
				ws.cell(row=rowIdx, column=2).number_format = _NUMBER_FMT
		ws.cell(row=rowIdx, column=3, value=unit)

	_autoWidth(ws)


def _writeDataFrameSheet(
	wb: Workbook,
	title: str,
	df: pl.DataFrame,
	*,
	columns: list[str] | None = None,
	maxRows: int = 0,
) -> None:
	if df.height == 0:
		return

	if columns:
		validCols = [c for c in columns if c in df.columns]
		if validCols:
			df = df.select(validCols)

	if maxRows > 0:
		df = df.head(maxRows)

	safeTitle = title[:31]
	existing = {ws.title for ws in wb.worksheets}
	if safeTitle in existing:
		safeTitle = safeTitle[:28] + "_2"

	ws = wb.create_sheet(title=safeTitle)
	_writeHeaders(ws, df.columns)

	for rowIdx, row in enumerate(df.iter_rows(named=True), 2):
		for colIdx, colName in enumerate(df.columns, 1):
			val = row[colName]
			if val is None:
				continue
			cell = ws.cell(row=rowIdx, column=colIdx, value=val)
			if isinstance(val, float):
				if abs(val) >= 1000:
					cell.number_format = _NEGATIVE_FMT
				else:
					cell.number_format = '0.00'
			elif isinstance(val, int):
				cell.number_format = _NUMBER_FMT

	_autoWidth(ws)
	ws.freeze_panes = "A2"


def _getAvailableModules(c: Company) -> list[tuple[str, str]]:
	from dartlab.company import _ALL_PROPERTIES
	available = []
	for name, label in _ALL_PROPERTIES:
		available.append((name, label))
	available.append(("ratios", "재무비율"))
	return available


def _resolveData(c: Company, moduleName: str) -> pl.DataFrame | None:
	if moduleName == "ratios":
		return None
	try:
		data = getattr(c, moduleName, None)
	except Exception:
		return None
	if isinstance(data, pl.DataFrame) and data.height > 0:
		return data
	return None


def exportToExcel(
	c: Company,
	outputPath: str | Path | None = None,
	modules: list[str] | None = None,
) -> str:
	"""Company 데이터를 .xlsx로 내보낸다.

	Args:
		c: Company 인스턴스.
		outputPath: 저장 경로. None이면 현재 디렉토리에 '{종목코드}_{회사명}.xlsx'.
		modules: 포함할 모듈 목록. None이면 데이터가 있는 모든 모듈.
			finance: "IS", "BS", "CF" (연도별 시계열 피벗)
			비율: "ratios"
			나머지: Company property 이름 (dividend, audit, employee, executive, ...)

	Returns:
		저장된 파일 경로 (str).
	"""
	allModules = _getAvailableModules(c)
	allNames = {name for name, _ in allModules}
	labelMap = {name: label for name, label in allModules}

	if modules is None:
		targetModules = [name for name, _ in allModules]
	else:
		targetModules = [m for m in modules if m in allNames]

	wb = Workbook()
	wb.remove(wb.active)

	financeModules = [m for m in targetModules if m in _FINANCE_SHEETS]
	if financeModules and c._hasFinance:
		from dartlab.engines.dart.finance.pivot import buildAnnual
		result = buildAnnual(c.stockCode)
		if result:
			series, years = result
			for sjDiv in financeModules:
				if sjDiv in series:
					_writeFinanceSheet(wb, sjDiv, series, years)

	if "ratios" in targetModules and c._hasFinance:
		_writeRatiosSheet(wb, c)

	for moduleName in targetModules:
		if moduleName in _FINANCE_SHEETS or moduleName in _SPECIAL_SHEETS:
			continue
		df = _resolveData(c, moduleName)
		if df is not None:
			label = labelMap.get(moduleName, moduleName)
			_writeDataFrameSheet(wb, label, df)

	if not wb.sheetnames:
		raise ValueError(f"{c.stockCode} ({c.corpName}): 내보낼 데이터 없음")

	if outputPath is None:
		safeName = c.corpName.replace("/", "_").replace("\\", "_")
		outputPath = Path.cwd() / f"{c.stockCode}_{safeName}.xlsx"
	else:
		outputPath = Path(outputPath)

	wb.save(str(outputPath))
	return str(outputPath)


def exportWithTemplate(
	c: Company,
	template: ExcelTemplate,
	outputPath: str | Path | None = None,
) -> str:
	"""템플릿 기반 Excel 내보내기.

	SheetSpec별로 소스를 로드하고, columns/years 필터를 적용하여 시트를 생성한다.

	Args:
		c: Company 인스턴스.
		template: ExcelTemplate (시트 구성 정의).
		outputPath: 저장 경로. None이면 자동 생성.

	Returns:
		저장된 파일 경로 (str).
	"""
	wb = Workbook()
	wb.remove(wb.active)

	annualCache: tuple[dict, list[str]] | None = None

	for spec in template.sheets:
		source = spec.source

		if source in _FINANCE_SHEETS:
			if not c._hasFinance:
				continue
			if annualCache is None:
				from dartlab.engines.dart.finance.pivot import buildAnnual
				result = buildAnnual(c.stockCode)
				if result is None:
					continue
				annualCache = result
			series, years = annualCache
			if source in series:
				_writeFinanceSheet(
					wb, source, series, years,
					label=spec.label,
					columns=spec.columns or None,
					yearFilter=spec.years or None,
				)
			continue

		if source == "ratios":
			if c._hasFinance:
				_writeRatiosSheet(wb, c, label=spec.label)
			continue

		df = _resolveData(c, source)
		if df is not None:
			_writeDataFrameSheet(
				wb, spec.label, df,
				columns=spec.columns or None,
				maxRows=spec.maxRows,
			)

	if not wb.sheetnames:
		raise ValueError(f"{c.stockCode} ({c.corpName}): 템플릿 '{template.name}'에 해당하는 데이터 없음")

	if outputPath is None:
		safeName = c.corpName.replace("/", "_").replace("\\", "_")
		templateSafe = template.name.replace("/", "_").replace("\\", "_")
		outputPath = Path.cwd() / f"{c.stockCode}_{safeName}_{templateSafe}.xlsx"
	else:
		outputPath = Path(outputPath)

	wb.save(str(outputPath))
	return str(outputPath)


def listAvailableModules(c: Company) -> list[dict[str, str]]:
	"""내보내기 가능한 모듈 목록 반환 (GUI/API용)."""
	result = []
	for name, label in _getAvailableModules(c):
		if name in _FINANCE_SHEETS:
			hasData = c._hasFinance
		elif name == "ratios":
			hasData = c._hasFinance
		else:
			hasData = _resolveData(c, name) is not None
		if hasData:
			result.append({"name": name, "label": label})
	return result
